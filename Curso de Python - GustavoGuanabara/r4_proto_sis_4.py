# -*- coding: utf-8 -*-
# ESSE PROGRAMA TEM O OBJETIVO DE CRIAR OS CÓDIGOS
# DOS EQUIPAMENTOS USADOS, CRIAR O CÓDIGO DAS ETIQUETAS
# E OS CÓDIGOS DOS RÓTULOS DE EQUIPAMENTOS
# OS RESULTADOS SERÃO CALCULADOS EM UMA PLANILHA
# DO EXCEL. (FUNCIONAMENTO: OK)

from openpyxl import load_workbook

arquivo = load_workbook("MAPA_APP_dev.xlsx")

mobil = arquivo["MOBILIZADO"]
apoio = arquivo["TABELAS_APOIO"]

Ulinha = mobil.max_row
apoio.cell(4, 5).value = "CÓDIGO"
inativo = 0
ativo = 0

for i in range(4,Ulinha+1):                 # GERA O CÓDIGO DO EQUIPAMENTO
    prefixo = mobil.cell(i, 5).value        # QUE SERÁ USADO NO MAPEAMENTO
    ordem = mobil.cell(i, 6).value
    propriedade = mobil.cell(i, 9).value
    mobilizado = mobil.cell(i, 4).value
    status = mobil.cell(i,11).value
    cod = f'{prefixo}-{ordem}'          # CONCATENAÇÃO PARA O CÓDIGO
    mobil.cell(i,7).value = cod
    apoio.cell(i+1, 5).value = cod

    etiqueta = f'{propriedade} / {cod}: {mobilizado}'   # ETIQUETA
    mobil.cell(i, 10).value = etiqueta
    if status == 'EXCLUÍDO' or status == 'DESLIGADO':
        rotulo = ''
        mobil.cell(i, 12).value = rotulo
        inativo += 1
        mobil.cell(1, 5).value = inativo
    else:
        rotulo = f'{cod} / {propriedade}'
        mobil.cell(i, 12).value = rotulo
        ativo += 1
        mobil.cell(2, 5).value = ativo

# OPERAÇÕES DE PESQUISA
# Preparar a pesquisa para ser consultada pelo COD do equipamento.
# 1. Ler toda a coluna de códigos e estabelecer uma consulta sem que os valores sejam repetidos.


arquivo.save("MAPA_APP_dev.xlsx")

