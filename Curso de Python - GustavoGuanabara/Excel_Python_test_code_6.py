from openpyxl import load_workbook


layout = [[sg.Text("Nome do pacote: ")],
          [sg.InputText()],
          [sg.Button("Ok")]]

window = sg.Window("Pesquisa de Dados", layout)

arquivo = load_workbook("Lista_de_pacotes_pip.xlsx")
entrada = arquivo["Planilha1"]
saida = arquivo["Planilha2"]

Ucoluna = entrada.max_column
Ulinha = entrada.max_row


for i in range(2, Ulinha + 1):
    package = entrada.cell(i, 1).value
    version = entrada.cell(i, 2).value
    event, valor = window.read()
    window.close()

    if package == event:
        saida.cell(3,1).value = package
        saida.cell(3,2).value  =version

sg.popup(f"Pesquisa de {values[0]} realizada com sucesso!")
arquivo.save("Lista_de_pacotes_pip_2.xlsx")