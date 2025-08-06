# NESSA ETAPA, O SISTEMA IRÁ IDENTIFICAR OS MESES DO FATURAMENTO
# E OS COLOCARÁ EM ORDEM CRESCENTE NA GUIA 'Análise'.

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo = load_workbook("CUSTOS_REPORT_1_GPT_091224.xlsx")

Base_Dados = arquivo["BASE"]
Resultados = arquivo["REPORT_1"]
DataRef_ini = Base_Dados.cell(3, 2).value
DataRef_fim = Base_Dados.cell(4, 2).value

# Última linha
Ulinha = Base_Dados.max_row

Base_Dados.cell(6, 2).value = Ulinha - 9  # O CABEÇALHO DA TABELA ESTÁ NA LINHA 9. SUBTRAIR 9 DO TOTAL DE LINHAS
Base_Dados.cell(6, 2).number_format = '#,##0'
Base_Dados.cell(6, 2).alignment = Alignment(horizontal='center', vertical='center')


for linha in range(10, Ulinha + 1):
    DataRef = Base_Dados.cell(linha, 3).value
    if DataRef >= DataRef_ini:
        if DataRef <= DataRef_fim:
            Resultados.cell(linha - 7, 1).value = DataRef
            Resultados.cell(linha - 7, 1).number_format = 'mm/yyyy'
            Resultados.cell(linha - 7, 1).alignment = Alignment(horizontal='center', vertical='center')
            print(DataRef)

arquivo.save("CUSTOS_REPORT_2_GPT_091224.xlsx")
