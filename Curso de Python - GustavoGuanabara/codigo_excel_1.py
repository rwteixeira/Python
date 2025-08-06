#import openpyxl
from openpyxl import load_workbook
arquivo = load_workbook("Alunos.xlsx")
# Ver as Guias
print(arquivo.sheetnames)
