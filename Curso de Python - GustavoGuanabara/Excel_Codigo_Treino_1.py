# OS SEGUINTES EXEMPLOS SÃO DIDÁTICOS.
# RICARDO WAGNER TEIXEIRA - 08/10/2024

from openpyxl import load_workbook
arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_atual = arquivo.active
#print(guia_atual)
guia_contas_pagar = arquivo["Python"]
# ultima linha
Ulinha = guia_contas_pagar.max_row
#print(Ulinha)
Ucoluna = guia_contas_pagar.max_column
#print(coluna)

#print(len(guia_contas_pagar["A"]))
#ultima_linha = len(guia_contas_pagar["A"])
#print(ultima_linha)
#ultima_coluna = len(guia_contas_pagar["1"])
#print(ultima_coluna)

# Atrasado, Confirmado, Não Pago
# Coluna 15 (Situação)

for linha in range(5, Ulinha + 1):
    situacao = guia_contas_pagar.cell(linha,11).value
    Valor = guia_contas_pagar.cell(linha, 12).value
    if situacao == "Atrasado":
        Total = (Valor * 1.3) + guia_contas_pagar.cell(linha, 14).value
        guia_contas_pagar.cell(linha, 15).value = 1.3
        guia_contas_pagar.cell(linha, 17).value = Total
    elif situacao == "Confirmado":
        Total = Valor
        guia_contas_pagar.cell(linha, 15).value = 0
        guia_contas_pagar.cell(linha, 17).value = Total
    else:
        situacao == "Não Pago"
        Total = Valor * 1.5 + guia_contas_pagar.cell(linha, 14).value
        guia_contas_pagar.cell(linha, 15).value = 1.5
        guia_contas_pagar.cell(linha, 17).value = Total

arquivo.save("CUSTO DA OPERACAO 3.xlsx")
