from openpyxl import load_workbook

# Carregar a planilha existente
arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")

# Nome das guias
guia_original = arquivo["Python"]

if "Análise" in arquivo.sheetnames:
    guia_totalizacao = arquivo["Análise"]
else:
    guia_totalizacao = arquivo.create_sheet(title="Análise")

# Criar cabeçalho da guia de totalização
for mes in range(1, 13):
    guia_totalizacao.cell(row=1, column=mes + 1, value=mes)  # Meses na linha de cabeçalho (2ª coluna em diante)

# Criar linhas para cada dia do mês (até o dia 31)
for dia in range(1, 32):
    guia_totalizacao.cell(row=dia + 1, column=1, value=dia)  # Dias na 1ª coluna (a partir da linha 2)

# Dicionário para armazenar total por dia e mês
totais = {}

# Percorrer as linhas da guia original e calcular os totais por dia e mês
for row in guia_original.iter_rows(min_row=5, values_only=True):
    mes = row[0]  # Supondo que o mês esteja na 1ª coluna
    dia = row[1]  # Supondo que o dia esteja na 2ª coluna
    valor = row[2]  # Supondo que o valor esteja na 3ª coluna

    if mes is not None and dia is not None and valor is not None:
        # Chave composta por dia e mês para acumular os valores
        chave = (dia, mes)
        if chave in totais:
            totais[chave] += valor
        else:
            totais[chave] = valor

# Preencher a guia de totalização com os valores calculados
for (dia, mes), total in totais.items():
    guia_totalizacao.cell(row=dia + 1, column=mes + 1, value=total)

# Salvar a planilha modificada
arquivo.save("CUSTO DA OPERACAO 3.xlsx")
print("Totalização concluída e salva em 'CUSTO DA OPERACAO 3.xlsx'.")
