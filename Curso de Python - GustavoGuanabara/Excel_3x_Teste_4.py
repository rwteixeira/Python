from openpyxl import load_workbook, Workbook

# Carregar a planilha Excel existente (ajuste o nome do arquivo e a folha conforme necessário)
wb = load_workbook("meses_ordenados.xlsx")
ws = wb["Planilha1"]

# Supondo que a coluna "A" contenha os números dos meses, e que a primeira linha é o cabeçalho
meses_numeros = []

# Ler todos os valores da coluna A, a partir da segunda linha, ignorando o cabeçalho
for cell in ws['B'][1:]:  # Começa a leitura a partir do índice 1
    if cell.value is not None:
        meses_numeros.append(cell.value)

# Obter os valores únicos e ordenar
meses_numeros_unicos = sorted(set(meses_numeros))

# Criar uma nova planilha para salvar os resultados
novo_wb = Workbook()
novo_ws = novo_wb.active
novo_ws.title = "Meses Ordenados Únicos"

# Escrever os meses únicos na nova planilha
for idx, mes in enumerate(meses_numeros_unicos, start=1):
    novo_ws.cell(row=idx, column=1, value=mes)

# Salvar em um novo arquivo Excel
novo_wb.save("meses_ordenados_unicos.xlsx")

print("Meses únicos ordenados e salvos em 'meses_ordenados_unicos.xlsx'.")
