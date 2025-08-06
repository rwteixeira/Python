# EXEMPLOS DE FORMATAÇÃO DA PLANILHA COM OPENPYXL
# FUNCIONA PERFEITAMENTE.

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create a workbook
workbook = Workbook()

# Get the active sheet
sheet = workbook.active

# Change content in cell
sheet['A1'] = "hello"
sheet['B1'] = "geek"

sheet.title = "Geek Sheet"

# Example: Tomato-colored, bold, and italic text
sheet['A1'].font = Font(
  					name='Calibri',
  					bold=True,
  					italic=True,
  					size=14,
  					color="FF6347"
				)


# Example: Solid yellow fill
sheet['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Change the allignment
sheet['C1'] = "I'm vertically and Horizontally Centered"
sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')


sheet.column_dimensions['A'].width = 10 # Diferença de 0,71 px

# Set border properties
border_thick = Side(style='thick')
border_thin = Side(style='thin')
borded_dashed = Side(style='dashed')
border_dotted = Side(style='dotted')

# Apply different border styles
sheet['C1'].border = Border(top=border_thick, left=border_thin, right=borded_dashed, bottom=border_dotted)


# Format number to 2 decimal places
sheet['A3'] = 4312345.6789
sheet['A4'] = 6
sheet['A3'].number_format = '#,##0.00'
sheet['A4'].number_format = '00'

# Format date
sheet['B3'] = "16-10-2024"
sheet['B3'].number_format = 'dd-mm-yyyy'

# Save the work book
workbook.save("hello_geek.xlsx")
