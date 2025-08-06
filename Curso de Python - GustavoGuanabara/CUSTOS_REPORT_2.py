from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

arquivo = load_workbook("CUSTOS_REPORT_1_GPT_091224.xlsx")

Base_Dados = arquivo["BASE"]
Resultados = arquivo["REPORT_1"]

DataRef_ini = Base_Dados["B3"].value
DataRef_fim = Base_Dados["B4"].value

data_inicio = DataRef_ini.strftime("%m/%Y")
data_fim = DataRef_fim.strftime("%m/%Y")

# Última linha
Ulinha = Base_Dados.max_row
Base_Dados.cell(6, 2).value = Ulinha - 9  # O CABEÇALHO DA TABELA ESTÁ NA LINHA 9. SUBTRAIR 9 DO TOTAL DE LINHAS
Base_Dados.cell(6, 2).number_format = '#,##0'
Base_Dados.cell(6, 2).alignment = Alignment(horizontal='center', vertical='center')

resultados = []

# Varre as linhas da planilha principal para realizar a consulta
for linha in Base_Dados.iter_rows(min_row=10, max_col=9, values_only=True):
    data_str = linha[2]  # Supondo que a data está na coluna C
    valor = linha[8]     # Supondo que o valor associado está na coluna B

    # Ignora linhas vazias
    if data_str is None:
        continue
    # Converte a data da planilha para datetime
    try:
        data_datetime = datetime(1899, 12, 30) + timedelta(days=data_str)
    except ValueError:
        continue  # Pula datas mal formatadas

    data_atual = data_datetime.strftime("%m/%Y")
    # Verifica se a data está no intervalo
    if data_inicio <= data_atual <= data_fim:
        resultados.append((data_str, valor))  # Adiciona à lista de resultados

# Acessa ou cria a Planilha2
if "REPORT_1" not in arquivo.sheetnames:
    planilha_resultados = arquivo.create_sheet("REPORT_1")
else:
    planilha_resultados = arquivo["REPORT_1"]

# Escreve os resultados na REPORT_1
planilha_resultados.append(["Data", "Valor"])  # Cabeçalho
for resultado in resultados:
    planilha_resultados.append(resultado)

# Salva o arquivo Excel com os resultados
arquivo.save("CUSTOS_REPORT_2_GPT.xlsx")
print("Consulta concluída e resultados salvos na guia REPORT_1!")
