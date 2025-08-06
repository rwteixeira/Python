from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_atual = arquivo.active

guia_contas_receber = arquivo["Python"]
guia_resultados = arquivo["Analise"]

Ulinha = guia_contas_receber.max_row # ultima linha
Ucoluna = guia_contas_receber.max_column # ultima coluna

dia = 1
mes = 8

while dia < 32:

    # Formatação do cabeçalho de coluna.
    guia_resultados.cell(1, 3).value = "Dia"
    guia_resultados.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    guia_resultados.cell(1, 3).font = Font(name='Consolas', size=11, bold=True)
    guia_resultados.cell(1, 4).value = "Mês " + str(mes)
    guia_resultados.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    guia_resultados.cell(1, 4).font = Font(name='Consolas', size=11, bold=True)

    # Formatação dos dias do calendário.
    guia_resultados.cell(dia + 1, 3).number_format = '00'
    guia_resultados.cell(dia + 1, 3).value = dia
    guia_resultados.cell(dia + 1, 3).alignment = Alignment(horizontal='center', vertical='center')
    dia += 1

# _____________________________________________________________
# NESSA ETAPA, O SISTEMA IRÁ IDENTIFICAR OS MESES DO FATURAMENTO
# E OS COLOCARÁ EM ORDEM CRESCENTE NA GUIA 'ANALISE'.

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_contas_receber = arquivo["Python"]
guia_resultados = arquivo["Analise"]

meses_numeros = []
guia_resultados.cell(1, 1).value = "Mês:"
guia_resultados.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
guia_resultados.cell(1, 1).font = Font(name='Consolas', size=11, bold=True)

for cell in guia_contas_receber['I'][4:]:
    if cell.value is not None:
        meses_numeros.append(cell.value)

meses_numeros_unicos = sorted(set(meses_numeros))
Resultado = meses_numeros_unicos

if "Analise" in arquivo.sheetnames:
    novo_ws = arquivo["Analise"]
else:
    novo_ws = arquivo.create_sheet(title="Analise")

for idx, mes in enumerate(Resultado, start=2):
    guia_resultados.cell(idx, 1).alignment = Alignment(horizontal='center', vertical='center')
    novo_ws.cell(row=idx, column=1, value=mes)

arquivo.save("CUSTO DA OPERACAO 3.xlsx")
print("Resultado:", Resultado)

# ______________________________________________________________
# OS SEGUINTES EXEMPLOS SÃO DIDÁTICOS.
# RICARDO WAGNER TEIXEIRA - 08/10/2024
# Ref: https://www.geeksforgeeks.org/merge-and-unmerge-excel-cells-using-openpyxl-in-python/?ref=oin_asr2

arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_atual = arquivo.active
# print(guia_atual)
guia_contas_receber = arquivo["Python"]

# ultima linha
Ulinha = guia_contas_receber.max_row
# print(Ulinha)
Ucoluna = guia_contas_receber.max_column

Total_Atrasado = 0
Total_NaoPago = 0
Total_Confirmado = 0

Total_Vl_Atrasado = 0
Total_Vl_Nao_Pago = 0
Total_Vl_Confirmado = 0

Total_Juros_A = 0
Total_Juros_N = 0
Total_Juros_C = 0

# ALTERANDO A LARGURA DAS COLUNAS DE RESUMO
guia_contas_receber.column_dimensions['S'].width = 18
guia_contas_receber.column_dimensions['T'].width = 18
guia_contas_receber.column_dimensions['U'].width = 18
guia_contas_receber.column_dimensions['V'].width = 18

guia_contas_receber.cell(4, 20).alignment = Alignment(horizontal='center', vertical='center')
guia_contas_receber.cell(4, 20).value = "Atrasado"
guia_contas_receber.cell(4, 20).font = Font(name='Consolas',size=11,bold=True)
guia_contas_receber.cell(4, 21).alignment = Alignment(horizontal='center', vertical='center')
guia_contas_receber.cell(4, 21).value = "Não Pago"
guia_contas_receber.cell(4, 21).font = Font(name='Consolas',size=11,bold=True)
guia_contas_receber.cell(4, 22).alignment = Alignment(horizontal='center', vertical='center')
guia_contas_receber.cell(4, 22).value = "Confirmado"
guia_contas_receber.cell(4, 22).font = Font(name='Consolas',size=11,bold=True)
guia_contas_receber.cell(4, 23).alignment = Alignment(horizontal='center', vertical='center')

guia_contas_receber.cell(5, 19).value = "Ocorrências:"
guia_contas_receber.cell(5, 19).font = Font(name='Consolas',size=11,bold=True)
guia_contas_receber.cell(6, 19).value = "Total:"
guia_contas_receber.cell(6, 19).font = Font(name='Consolas',size=11,bold=True)
guia_contas_receber.cell(7, 19).value = "Juros:"
guia_contas_receber.cell(7, 19).font = Font(name='Consolas',size=11,bold=True)


for linha in range(5, Ulinha + 1):
    situacao = guia_contas_receber.cell(linha,11).value
    juros = guia_contas_receber.cell(linha, 14).value
    valor = guia_contas_receber.cell(linha, 12).value
    taxa = guia_contas_receber.cell(linha, 15).value
    desconto = guia_contas_receber.cell(linha, 13).value

    if situacao == "Atrasado":
        if juros == 0:
            mora = 1.3
            Total = valor * mora
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(linha, 17).number_format = '#,##0.00'
            guia_contas_receber.cell(5, 20).value = Total_Atrasado = Total_Atrasado + 1
            guia_contas_receber.cell(5, 20).alignment = Alignment(horizontal='center', vertical='center')
            Total_Vl_Atrasado += Total
            guia_contas_receber.cell(6, 20).value = Total_Vl_Atrasado
            guia_contas_receber.cell(6, 20).number_format = '#,##0.00'
            guia_contas_receber.cell(6, 20).alignment = Alignment(horizontal='center', vertical='center')
        else:
            mora = 1.3
            Total = valor * mora + juros
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(5, 20).value = Total_Atrasado = Total_Atrasado + 1
            guia_contas_receber.cell(5, 20).alignment = Alignment(horizontal='center', vertical='center')
            Total_Vl_Atrasado += Total
            guia_contas_receber.cell(6, 20).value = Total_Vl_Atrasado
            guia_contas_receber.cell(6, 20).number_format = '#,##0.00'
            guia_contas_receber.cell(6, 20).alignment = Alignment(horizontal='center', vertical='center')
            Total_Juros_A += juros
            guia_contas_receber.cell(7, 20).value = Total_Juros_A
            guia_contas_receber.cell(7, 20).number_format = '#,##0.00'
            guia_contas_receber.cell(7, 20).alignment = Alignment(horizontal='center', vertical='center')
    elif situacao == "Não Pago":
        if juros == 0:
            mora = 1.5
            Total = valor * mora
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(5, 21).value = Total_NaoPago = Total_NaoPago + 1
            guia_contas_receber.cell(5, 21).alignment = Alignment(horizontal='center', vertical='center')
            Total_Vl_Nao_Pago += Total
            guia_contas_receber.cell(6, 21).value = Total_Vl_Nao_Pago
            guia_contas_receber.cell(6, 21).number_format = '#,##0.00'
            guia_contas_receber.cell(6, 21).alignment = Alignment(horizontal='center', vertical='center')
        else:
            mora = 1.5
            Total = valor * mora + juros
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(5, 21).value = Total_NaoPago = Total_NaoPago + 1
            guia_contas_receber.cell(5, 21).alignment = Alignment(horizontal='center', vertical='center')
            Total_Vl_Nao_Pago += Total
            guia_contas_receber.cell(6, 21).value = Total_Vl_Nao_Pago
            Total_Juros_N += juros
            guia_contas_receber.cell(7, 21).value = Total_Juros_N
            guia_contas_receber.cell(7, 21).number_format = '#,##0.00'
            guia_contas_receber.cell(7, 21).alignment = Alignment(horizontal='center', vertical='center')
    elif situacao == "Confirmado":
        if juros == 0:
            mora = 0
            desc = 0.035
            Total = valor - (valor * desc)
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 13).value = desc
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(5, 22).value = Total_Confirmado = Total_Confirmado + 1
            guia_contas_receber.cell(5, 22).alignment = Alignment(horizontal='center', vertical='center')
            Total_Vl_Confirmado += Total
            guia_contas_receber.cell(6, 22).value = Total_Vl_Confirmado
            guia_contas_receber.cell(6, 22).number_format = '#,##0.00'
            guia_contas_receber.cell(6, 22).alignment = Alignment(horizontal='center', vertical='center')
        else:
            mora = 0
            desc = 0.035
            Total = valor + juros
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 13).value = desc
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(5, 22).value = Total_Confirmado = Total_Confirmado + 1
            guia_contas_receber.cell(5, 22).alignment = Alignment(horizontal='center', vertical='center')
            Total_Vl_Confirmado += Total
            guia_contas_receber.cell(6, 22).value = Total_Vl_Confirmado
            guia_contas_receber.cell(6, 22).number_format = '#,##0.00'
            guia_contas_receber.cell(6, 22).alignment = Alignment(horizontal='center', vertical='center')
            Total_Juros_C += juros
            guia_contas_receber.cell(7, 22).value = Total_Juros_C
            guia_contas_receber.cell(7, 22).number_format = '#,##0.00'
            guia_contas_receber.cell(7, 22).alignment = Alignment(horizontal='center', vertical='center')

arquivo.save("CUSTO DA OPERACAO 3.xlsx")
