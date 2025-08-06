# OBJETIVO: ALTERAR A TAXA DE JUROS PARA CÁLCULO DO TOTAL DE 1,2 PARA 1,3;
#           CALCULAR O VALOR TOTAL COM A NOVA TAXA;
#           SALVAR O ARQUIVO COM O NOME "Contas_Receber_OPENPYXL.xlsx"
# Ref: https://www.youtube.com/watch?v=IT7zPluDADk

# import openpyxl
from openpyxl import workbook, load_workbook

planilha = load_workbook("ContasReceber1.xlsx")
aba_ativa = planilha.active

for celula in aba_ativa["J"]:
    if celula.value == "Atrasado":
        linha = celula.row
        aba_ativa[f"M{linha}"] = 1.3
        valor = float(aba_ativa.cell(linha, 11).value)
        taxa = float(aba_ativa.cell(linha, 13).value)
        juros = float(aba_ativa.cell(linha, 12).value)
        if juros == 0:
            total = valor * taxa
        else:
            total = (valor * taxa) + juros
        aba_ativa.cell(linha, 14).value = total
    elif celula.value == "Não Pago":
        linha = celula.row
        aba_ativa[f"M{linha}"] = 1.5
        valor = float(aba_ativa.cell(linha, 11).value)
        taxa = float(aba_ativa.cell(linha, 13).value)
        juros = float(aba_ativa.cell(linha, 12).value)
        if juros == 0:
            total = valor * taxa
        else:
            total = (valor * taxa) + juros
        aba_ativa.cell(linha, 14).value = total
    elif celula.value == "Confirmado":
        linha = celula.row
        aba_ativa[f"M{linha}"] = 1
        valor = float(aba_ativa.cell(linha, 11).value)
        taxa = float(aba_ativa.cell(linha, 13).value)
        juros = float(aba_ativa.cell(linha, 12).value)
        if juros == 0:
            total = valor * taxa
        else:
            total = (valor * taxa) + juros
        aba_ativa.cell(linha, 14).value = total


planilha.save("CONTAS_RECEBER_openpyxl.xlsx")
