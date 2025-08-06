import openpyxl
#
# CARREGA UMA PLANILHA EXISTENTE.
# 
# 02/10/2024

# Carregando arquivo
book = openpyxl.load_workbook('ContasReceber2.xlsx')

# Selecionando uma página
guia = book['Planilha1']

# Imprimindo os dados de cada linha
# for rows in guia.iter_rows(min_row=2, max_row=10):  # Sem parâmetros, vai a tabela toda.
for rows in guia.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        #print(cell.value)
        print(f'{rows[1].value},{rows[10].value}')
# Faz alterações na planilha        
#        if cell.value == '75365':
#            cell.value = '11111'

# Salvar as alterações
book.save('ContasReceber2.xlsx')
