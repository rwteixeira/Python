# [Em desenvolvimento - 21/12/2024]

from openpyxl import load_workbook
from openpyxl.styles import Alignment
#from openpyxl.utils.datetime import from_excel
#from datetime import datetime, timedelta

# Carrega o arquivo Excel
arquivo = load_workbook("CUSTOS_REPORT_1_GPT_091224.xlsx")

# Acessa as planilhas
Base_Dados = arquivo["BASE"]

# Lê os campos de consulta
#DataRef_ini = Base_Dados["B3"].value
#DataRef_fim = Base_Dados["B4"].value
#Entid = Base_Dados["E3"].value
#Descr = Base_Dados["E4"].value

Plano_de_contas = Base_Dados["E5"].value

# Última linha
Ulinha = Base_Dados.max_row
Base_Dados.cell(6, 2).value = Ulinha - 9  # O cabeçalho da tabela está na linha 9
Base_Dados.cell(6, 2).number_format = '#,##0'
Base_Dados.cell(6, 2).alignment = Alignment(horizontal='center', vertical='center')

# Lista para armazenar os resultados
resultados = []


# Salva o arquivo Excel com os resultados
arquivo.save("CUSTOS_REPORT_3_GPT.xlsx")

# ----------------------------------------------------------
from openpyxl import load_workbook

# Abrir o arquivo Excel
arquivo = load_workbook("dados_financeiros.xlsx")

# Selecionar as guias
guia_origem = arquivo["Base"]  # Substituir pelo nome da guia de origem
guia_destino = arquivo["Report"]  # Substituir pelo nome da guia de destino

# Ler o critério de consulta na célula E5
plano_de_contas = guia_origem["E5"].value

# Verificar se o critério não está vazio
if not plano_de_contas:
    raise ValueError("O critério de consulta (célula E5) está vazio.")

# Obter os dados da guia de origem (supondo que começam na linha 10, com cabeçalho na linha 9)
dados_filtrados = []
for linha in guia_origem.iter_rows(min_row=10, max_row=guia_origem.max_row, values_only=True):
    if linha and plano_de_contas in linha:  # Verifica se o critério está presente na linha
        dados_filtrados.append(linha)

# Escrever os dados filtrados na guia de destino
linha_destino = 2  # Começar na célula B2
for dado in dados_filtrados:
    for col, valor in enumerate(dado, start=2):  # Começa a partir da coluna B
        guia_destino.cell(row=linha_destino, column=col, value=valor)
    linha_destino += 1

# Salvar o arquivo Excel com os resultados
arquivo.save("dados_financeiros_resultado.xlsx")
print("Consulta concluída! Os dados foram salvos na guia 'Report'.")

