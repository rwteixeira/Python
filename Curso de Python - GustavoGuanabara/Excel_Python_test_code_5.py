#
import logging
from openpyxl import load_workbook
from openpyxl import Workbook

def pasta_de_trabalho():
    nome_arquivo = str(input('Defina o nome do arquivo: '))
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet(title="Planilha1")
    ws2 = wb.create_sheet(title="Planilha2")
    wb.save(nome_arquivo + '.xlsx')

# Definir o nome da Pasta de trabalho do Excel
nome_do_arquivo = str(input('Digite o nome da pasta de trabalho: '))

# Configuração do logger
logging.basicConfig(
    filename='Excel Python 1.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

try:

    arquivo = load_workbook(nome_do_arquivo + '.xlsx')

    logging.info(f"Arquivo {arquivo} + '.xlsx' criado com sucesso.")

    # Carregar o arquivo Excel
    arquivo = load_workbook(nome_do_arquivo + '.xlsx')
    logging.info(f"Arquivo Excel {nome_do_arquivo} + '.xlsx' carregado com sucesso.")

    # Lista de guias padrão a serem removidas
    GuiasPadrão = ['Planilha1', 'Planilha2', 'Planilha3', 'Plan1', 'Plan2', 'Plan3', 'Sheet']

    # Iterar sobre as guias padrão e removê-las se existirem no arquivo
    for guia in GuiasPadrão:
        if guia in arquivo.sheetnames:
            planilha_para_remover = arquivo[guia]
            arquivo.remove(planilha_para_remover)
            logging.info(f'Planilha "{guia}" removida com sucesso.')
            # print(f'A guia "{guia}" foi removida com sucesso.')
        else:
            logging.warning(f'Planilha "{guia}" não encontrada para remoção.')
            # print(f'A guia "{guia}" não existe no arquivo.')

    # Criar as guias 'Entradas' e 'Relatórios' se não existirem
    if "Entradas" not in arquivo.sheetnames:
        arquivo.create_sheet("Entradas")
        logging.info('Planilha "Entradas" criada com sucesso.')
        # print('A guia "Entradas" foi criada.')
    else:
        logging.info('Planilha "Entradas" já existe.')
        # print('A guia "Entradas" já existe.')

    if "Relatórios" not in arquivo.sheetnames:
        arquivo.create_sheet("Relatórios")
        logging.info('Planilha "Relatórios" criada com sucesso.')
        # print('A guia "Relatórios" foi criada.')
    else:
        logging.info('Planilha "Relatórios" já existe.')
        # print('A guia "Relatórios" já existe.')

    # Salvar as alterações no arquivo Excel

    arquivo.save(arquivo + '.xlsx')
    logging.info(f"Arquivo Excel {arquivo} + '.xlsx' salvo com sucesso.")
    # print('As alterações foram salvas no arquivo "Excel_Python_treino_1.xlsx".')

except Exception as e:
    logging.error(f'Ocorreu um erro: {e}')
