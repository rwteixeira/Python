# -*- coding: utf-8 -*-
# COMPARA DUAS COLUNAS: 'A' e 'B' E EXIBE NUMA CAIXA DE MENSAGEM OS CÓDIGOS QUE REPETEM.
from openpyxl import load_workbook

# Carregar a planilha
file_path = "MAPA_APP_tabelas_1.xlsx"  # Atualize com o caminho correto
workbook = load_workbook(file_path)
sheet = workbook.active  # Usar a aba ativa


# Coletar os valores das colunas A e B a partir da linha 3
coluna_A = {row[0] for row in sheet.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True) if row[0]}
coluna_B = {row[0] for row in sheet.iter_rows(min_row=3, min_col=2, max_col=2, values_only=True) if row[0]}


# Encontrar valores que estão em A mas, não em B
valores_exclusivos_A = coluna_A.difference(coluna_B)


# Exibir a lista dos valores encontrados
print("Valores que estão na coluna A mas não estão na coluna B:")
#print(valores_exclusivos_A)

for valor, cod in enumerate(valores_exclusivos_A):
    #print(f'{valor} - {cod}')
    print(valores_exclusivos_A[valor])