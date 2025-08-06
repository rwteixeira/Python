# -*- coding: utf-8 -*-
# ESSE EXEMPLO DE CÓDIGO PODE SER CONSIDERADO FINAL.
# 30/01/2025 - BY: RWT
# O programa vai abrir uma pasta do Excel e verificar se as guias 'Planilha1, Planilha2 e 3 existem.
# Se existirem, serão apagadas.  Serão criadas duas outras guias: 'Entradas' e 'Relatórios', se não existirem.
# FUNCIONA PERFEITAMENTE.

import logging
from openpyxl import load_workbook
import easygui

# Configuração do logger
logging.basicConfig(
    filename='alteracoes.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

opcoes = ["Novo", "Abrir", "Remover"]
selecionado = easygui.choicebox("Opção:", title="Arquivos", choices=opcoes)
easygui.msgbox(f"Sua Opção: {selecionado}", title="Operação")

try:
    # Carregar o arquivo Excel
    arquivo = load_workbook(easygui.fileopenbox(title="Selecione um Arquivo"))
    logging.info(f'Arquivo {arquivo} carregado com sucesso.')


    # Lista de guias padrão a serem removidas
    GuiasPadrão = ['Planilha1', 'Planilha2', 'Planilha3', 'Plan1', 'Plan2', 'Plan3']

    # Iterar sobre as guias padrão e removê-las se existirem no arquivo
    for guia in GuiasPadrão:
        if guia in arquivo.sheetnames:
            planilha_para_remover = arquivo[guia]
            arquivo.remove(planilha_para_remover)
            logging.info(f'Planilha "{guia}" removida com sucesso.')
            # print(f'A guia "{guia}" foi removida com sucesso.')
        else:
            logging.warning(f'Planilha "{guia}" não encontrada para remoção.')
            # print(f'A guia "{guia}" não existe no arquivo.')

    # Criar as guias 'Entradas' e 'Relatórios' se não existirem
    if "Entradas" not in arquivo.sheetnames:
        arquivo.create_sheet("Entradas")
        logging.info('Planilha "Entradas" criada com sucesso.')
        # print('A guia "Entradas" foi criada.')
    else:
        logging.info('Planilha "Entradas" já existe.')
        # print('A guia "Entradas" já existe.')

    if "Relatórios" not in arquivo.sheetnames:
        arquivo.create_sheet("Relatórios")
        logging.info('Planilha "Relatórios" criada com sucesso.')
        # print('A guia "Relatórios" foi criada.')
    else:
        logging.info('Planilha "Relatórios" já existe.')
        # print('A guia "Relatórios" já existe.')

    # Salvar as alterações no arquivo Excel
    arquivo.save("Excel_Python_treino_X.xlsx")
    logging.info('Arquivo Excel salvo como "Excel_Python_treino_1.xlsx".')
    # print('As alterações foram salvas no arquivo "Excel_Python_treino_1.xlsx".')

except Exception as e:
    logging.error(f'Ocorreu um erro: {e}')
