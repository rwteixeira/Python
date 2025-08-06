from openpyxl import load_workbook

#arquivo = load_workbook("Alunos.xlsx")
arquivo = load_workbook("ContasReceber1.xlsx")

# ver as abas
print(arquivo.sheetnames)

# pegar a aba ativa
aba_atual = arquivo.active
print(aba_atual)

# selecionar uma aba específica
guia_contas_receber = arquivo["ContasReceber"]
print(guia_contas_receber)

# selecionar células
valor_a1 = guia_contas_receber["A1"].value
valor_b1 = guia_contas_receber.cell(row=1, column=2).value
print(valor_b1)

guia_contas_receber.cell(row=1, column=2).value = "PROVA 1"

arquivo.save("ContasReceber2.xlsx")

# ultima linha
print(guia_contas_receber.max_row)
print(len(guia_contas_receber["A"]))
