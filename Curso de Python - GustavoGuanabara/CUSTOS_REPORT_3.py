# [Última versão estável - 17/12/2024]
# Essa etapa está concluída.

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.datetime import from_excel
from datetime import datetime, timedelta

# Carrega o arquivo Excel
arquivo = load_workbook("CUSTOS_REPORT_1_GPT_091224.xlsx")

# Acessa as planilhas
Base_Dados = arquivo["BASE"]

# Lê as datas de referência
DataRef_ini = Base_Dados["B3"].value
DataRef_fim = Base_Dados["B4"].value
Entid = Base_Dados["E3"].value
Descr = Base_Dados["E4"].value

# Verifica se os campos de DataRef estão vazios
if DataRef_ini == None:
    if DataRef_fim == None:
        print("Consulta vazia. É preciso estabelecer as datas de inínio e fim!")
        exit()

# Certifica-se de que as datas são objetos datetime
if isinstance(DataRef_ini, datetime):
    data_inicio = DataRef_ini
elif isinstance(DataRef_ini, int) or isinstance(DataRef_ini, float):
    data_inicio = from_excel(DataRef_ini)  # Se for serial
else:
    data_inicio = datetime.strptime(DataRef_ini, "%m/%Y")  # Se for string

if isinstance(DataRef_fim, datetime):
    data_fim = DataRef_fim
elif isinstance(DataRef_fim, int) or isinstance(DataRef_fim, float):
    data_fim = from_excel(DataRef_fim)  # Se for serial
else:
    data_fim = datetime.strptime(DataRef_fim, "%m/%Y")  # Se for string

# Última linha
Ulinha = Base_Dados.max_row
Base_Dados.cell(6, 2).value = Ulinha - 9  # O cabeçalho da tabela está na linha 9
Base_Dados.cell(6, 2).number_format = '#,##0'
Base_Dados.cell(6, 2).alignment = Alignment(horizontal='center', vertical='center')

# Lista para armazenar os resultados
resultados = []

# Itera pelas linhas da planilha para realizar a consulta
for linha in Base_Dados.iter_rows(min_row=10, max_col=9, values_only=True):
    data_serial = linha[2]  # Supondo que a data está na coluna C
    valor = linha[8]        # Supondo que o valor está na coluna I

    # Ignora linhas vazias
    if data_serial is None:
        continue

    # Converte a data da planilha para datetime
    if isinstance(data_serial, datetime):
        data_datetime = data_serial
    elif isinstance(data_serial, int) or isinstance(data_serial, float):
        data_datetime = from_excel(data_serial)
    else:
        continue  # Ignora valores mal formatados

    # Verifica se a data está no intervalo
    if data_inicio <= data_datetime <= data_fim:
        resultados.append((data_datetime.strftime("%m/%Y"), valor))  # Formata para o padrão esperado

# Cria a planilha de resultados, se necessário
if "REPORT_1" not in arquivo.sheetnames:
    planilha_resultados = arquivo.create_sheet("REPORT_1")
else:
    planilha_resultados = arquivo["REPORT_1"]

# Escreve os resultados na planilha "REPORT_1"
planilha_resultados.append(["Data", "Valor"])  # Cabeçalho
for resultado in resultados:
    planilha_resultados.append(resultado)

# Salva o arquivo Excel com os resultados
arquivo.save("CUSTOS_REPORT_2_GPT.xlsx")
print(f"Consulta concluída. {len(resultados)} registros foram salvos na guia REPORT_1.")
