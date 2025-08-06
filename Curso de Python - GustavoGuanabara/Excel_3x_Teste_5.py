from openpyxl import load_workbook, Workbook

# Carregar a planilha Excel existente (ajuste o nome do arquivo e a folha conforme necessário)
wb: Workbook = load_workbook("CUSTO DA OPERACAO 2.xlsx")
ws = wb["Python"]

# Supondo que a coluna "B" contenha os números dos meses, e que a primeira linha é o cabeçalho
meses_numeros = []

# Ler todos os valores da coluna B, a partir da segunda linha, ignorando o cabeçalho
for cell in ws['I'][4:]:  # Começa a leitura a partir do índice 4 (inicia na 5ª linha, pois, começa em 0)
    if cell.value is not None:
        meses_numeros.append(cell.value)

# Obter os valores únicos e ordenar
meses_numeros_unicos = sorted(set(meses_numeros))

# Armazenar o resultado numa lista chamada 'Resultado'
Resultado = meses_numeros_unicos

# Criar uma nova planilha para salvar os resultados
#novo_wb = Workbook()
#novo_ws = novo_wb.active
#novo_ws.title = "Meses Ordenados Únicos"

if "Resultado" in wb.sheetnames:
    novo_ws = wb["Resultado"]
else:
    novo_ws = wb.create_sheet(title="Resultado")

# Escrever os meses únicos na nova planilha
for idx, mes in enumerate(Resultado, start=1):
    novo_ws.cell(row=idx, column=1, value=mes)

# Salvar em um novo arquivo Excel
wb.save("CUSTO DA OPERACAO 3.xlsx")



print("Meses únicos ordenados e salvos em 'CUSTO DA OPERACAO 3.xlsx'.")
print("Resultado:", Resultado)

