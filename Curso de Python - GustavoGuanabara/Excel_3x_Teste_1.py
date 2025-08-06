from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_atual = arquivo.active
#print(guia_atual)
guia_contas_receber = arquivo["Python"]
guia_resultados = arquivo["Analise"]
# ultima linha
Ulinha = guia_contas_receber.max_row
# print(Ulinha)
Ucoluna = guia_contas_receber.max_column

# -----------------------------------------------
dia = 1
mes = 8

while dia < 32:

    # Formatação do cabeçalho de coluna.
    guia_resultados.cell(1, 3).value = "Dia"
    guia_resultados.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    guia_resultados.cell(1, 3).font = Font(name='Consolas', size=11, bold=True)
    guia_resultados.cell(1, 4).value = "Mês " + str(mes)
    guia_resultados.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    guia_resultados.cell(1, 4).font = Font(name='Consolas', size=11, bold=True)

    # Formatação dos dias do calendário.
    guia_resultados.cell(dia + 1, 3).number_format = '00'
    guia_resultados.cell(dia + 1, 3).value = dia
    guia_resultados.cell(dia + 1, 3).alignment = Alignment(horizontal='center', vertical='center')
    dia += 1

# -----------------------------------------------


arquivo.save("CUSTO DA OPERACAO 3.xlsx")
