import openpyxl
from openpyxl import load_workbook

# Criar uma planilha (book)
book = openpyxl.Workbook()

# Como visualizar páginas existentes
print(book.sheetnames)

# Como criar uma página
book.create_sheet('Ativos')

# Como selecionar uma página
Ativos_page = book['Ativos']
Ativos_page.append(['Descrição', 'Qtde', 'Preço'])
Ativos_page.append(['Servidores', '12', 'R$ 30000,00'])
Ativos_page.append(['Roteadores', '3', 'R$ 1500,00'])
Ativos_page.append(['Storages', '1', 'R$ 35000,00'])
Ativos_page.append(['UPS', '2', 'R$ 90000,00'])

# Salvar a planilha
book.save('Panilha de orcamento.xlsx')
