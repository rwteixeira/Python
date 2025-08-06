# OS SEGUINTES EXEMPLOS SÃO DIDÁTICOS.
# RICARDO WAGNER TEIXEIRA - 08/10/2024

from openpyxl import load_workbook
arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_atual = arquivo.active
#print(guia_atual)
guia_contas_receber = arquivo["Python"]
# ultima linha
Ulinha = guia_contas_receber.max_row
#print(Ulinha)
Ucoluna = guia_contas_receber.max_column
#print(Ucoluna)

#print(len(guia_contas_receber["A"]))
#ultima_linha = len(guia_contas_receber["A"])
#print(ultima_linha)
#ultima_coluna = len(guia_contas_receber["1"])
#print(ultima_coluna)

# Atrasado, Confirmado, Não Pago
# Coluna 15 (Situação)
#

for linha in range(5, Ulinha + 1):
    situacao = guia_contas_receber.cell(linha,11).value
    juros = guia_contas_receber.cell(linha, 14).value
    valor = guia_contas_receber.cell(linha, 12).value
    taxa = guia_contas_receber.cell(linha, 15).value
    desconto = guia_contas_receber.cell(linha, 13).value
    Valor_Total = guia_contas_receber.cell(linha, 17).value

    if situacao == "Atrasado":
        if juros == 0:
            total = valor * 1.3
            taxa = 1.3
            total = Valor_Total
        else:
            total = valor * 1.3 + juros
            taxa = 1.3
            total = Valor_Total
    elif situacao == "Não Pago":
        if juros == 0:
            total = valor * 1.5
            taxa = 1.5
            total = Valor_Total
        else:
            total = valor * 1.5 + juros
            taxa = 1.5
            total = Valor_Total
    elif situacao == "Confirmado":
        if juros == 0:
            taxa = 0
            total = valor-(valor*0.035)
            desconto = 0.035
            total = Valor_Total
        else:
            total = valor + juros
            total = Valor_Total

arquivo.save("CUSTO DA OPERACAO 3.xlsx")
