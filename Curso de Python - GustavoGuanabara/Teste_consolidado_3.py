# NESSA ETAPA, O SISTEMA IRÁ IDENTIFICAR OS MESES DO FATURAMENTO
# E OS COLOCARÁ EM ORDEM CRESCENTE NA GUIA 'Análise'.

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_contas_receber = arquivo["Python"]
guia_resultados = arquivo["Análise"]

meses_numeros = []
guia_resultados.cell(1, 1).value = "Meses:"
guia_resultados.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
guia_resultados.cell(1, 1).font = Font(name='Consolas', size=11, bold=True)

for cell in guia_contas_receber['I'][4:]:
    if cell.value is not None:
        meses_numeros.append(cell.value)

meses_numeros_unicos = sorted(set(meses_numeros))
Meses_Relat = meses_numeros_unicos

if "Análise" in arquivo.sheetnames:
    novo_ws = arquivo["Análise"]
else:
    novo_ws = arquivo.create_sheet(title="Análise")

for idx, mes in enumerate(Meses_Relat, start=2):
    guia_resultados.cell(idx, 1).alignment = Alignment(horizontal='center', vertical='center')
    novo_ws.cell(row=idx, column=1, value=mes)
# ___________________________________
# COM BASE EM 'RESULTADO' O LAÇO FOR CRIARÁ OS CABEÇALHOS DE COLUNA
# DE DETALHAMENTO MENSAL.
ContMeses = len(Meses_Relat)

for MesCol in range(1, ContMeses + 1):
    guia_resultados.cell(1, 3 + MesCol).value = "Mês " + str(Meses_Relat[MesCol - 1])
    guia_resultados.cell(1, 3 + MesCol).alignment = Alignment(horizontal='center', vertical='center')
    guia_resultados.cell(1, 3 + MesCol).font = Font(name='Consolas', size=11, bold=True)
# ___________________________________
dia = 1
while dia < 32:
    # Formatação do cabeçalho de coluna.
    guia_resultados.cell(1, 3).value = "Dia"
    guia_resultados.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    guia_resultados.cell(1, 3).font = Font(name='Consolas', size=11, bold=True)

    # Formatação dos dias do calendário.
    guia_resultados.cell(dia + 1, 3).number_format = '00'
    guia_resultados.cell(dia + 1, 3).value = dia
    guia_resultados.cell(dia + 1, 3).alignment = Alignment(horizontal='center', vertical='center')
    dia += 1
# ___________________________________

# ultima linha
Ulinha = guia_contas_receber.max_row
# print(Ulinha)
Ucoluna = guia_contas_receber.max_column

Total_Atrasado = 0
Total_NaoPago = 0
Total_Confirmado = 0

Total_Vl_Atrasado = 0
Total_Vl_Nao_Pago = 0
Total_Vl_Confirmado = 0

Total_Juros_A = 0
Total_Juros_N = 0
Total_Juros_C = 0

# ALTERANDO A LARGURA DAS COLUNAS DE RESUMO
guia_resultados.column_dimensions['I'].width = 18
guia_resultados.column_dimensions['J'].width = 18
guia_resultados.column_dimensions['K'].width = 18
guia_resultados.column_dimensions['L'].width = 18

guia_resultados.cell(1, 10).value = "Atrasado"
guia_resultados.cell(1, 10).alignment = Alignment(horizontal='center', vertical='center')
guia_resultados.cell(1, 10).font = Font(name='Consolas', size=11, bold=True)

guia_resultados.cell(1, 11).value = "Não Pago"
guia_resultados.cell(1, 11).alignment = Alignment(horizontal='center', vertical='center')
guia_resultados.cell(1, 11).font = Font(name='Consolas', size=11, bold=True)

guia_resultados.cell(1, 12).value = "Confirmado"
guia_resultados.cell(1, 12).alignment = Alignment(horizontal='center', vertical='center')
guia_resultados.cell(1, 12).font = Font(name='Consolas', size=11, bold=True)

guia_resultados.cell(2, 9).value = "Ocorrências:"
guia_resultados.cell(2, 9).font = Font(name='Consolas', size=11, bold=True)

guia_resultados.cell(3, 9).value = "Juros:"
guia_resultados.cell(3, 9).font = Font(name='Consolas', size=11, bold=True)

guia_resultados.cell(4, 9).value = "Total:"
guia_resultados.cell(4, 9).font = Font(name='Consolas', size=11, bold=True)

for linha in range(5, Ulinha + 1):
    situacao = guia_contas_receber.cell(linha, 11).value
    juros = guia_contas_receber.cell(linha, 14).value
    valor = guia_contas_receber.cell(linha, 12).value
    taxa = guia_contas_receber.cell(linha, 15).value
    desconto = guia_contas_receber.cell(linha, 13).value

    if situacao == "Atrasado":
        if juros == 0:
            mora = 1.3
            Total = valor * mora
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total
            guia_contas_receber.cell(linha, 17).number_format = '#,##0.00'

            ocorrências = Total_Atrasado = Total_Atrasado + 1
            guia_resultados.cell(2, 10).value = ocorrências
            guia_resultados.cell(2, 10).alignment = Alignment(horizontal='center', vertical='center')

            Total_Vl_Atrasado += Total
            guia_resultados.cell(4, 10).value = Total_Vl_Atrasado
            guia_resultados.cell(4, 10).number_format = '#,##0.00'
            guia_resultados.cell(4, 10).alignment = Alignment(horizontal='center', vertical='center')
        else:
            mora = 1.3
            Total = valor * mora + juros
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total

            ocorrências = Total_Atrasado = Total_Atrasado + 1
            guia_resultados.cell(2, 10).value = ocorrências
            guia_resultados.cell(2, 10).alignment = Alignment(horizontal='center', vertical='center')

            Total_Vl_Atrasado += Total
            guia_resultados.cell(4, 10).value = Total_Vl_Atrasado
            guia_resultados.cell(4, 10).number_format = '#,##0.00'
            guia_resultados.cell(4, 10).alignment = Alignment(horizontal='center', vertical='center')

            Total_Juros_A += juros
            guia_resultados.cell(3, 10).value = Total_Juros_A
            guia_resultados.cell(3, 10).number_format = '#,##0.00'
            guia_resultados.cell(3, 10).alignment = Alignment(horizontal='center', vertical='center')
    elif situacao == "Não Pago":
        if juros == 0:
            mora = 1.5
            Total = valor * mora
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total

            ocorrências = Total_NaoPago = Total_NaoPago + 1
            guia_resultados.cell(2, 11).value = ocorrências
            guia_resultados.cell(2, 11).alignment = Alignment(horizontal='center', vertical='center')

            Total_Vl_Nao_Pago += Total
            guia_resultados.cell(4, 11).value = Total_Vl_Nao_Pago
            guia_resultados.cell(4, 11).number_format = '#,##0.00'
            guia_resultados.cell(4, 11).alignment = Alignment(horizontal='center', vertical='center')
        else:
            mora = 1.5
            Total = valor * mora + juros
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 17).value = Total

            ocorrências = Total_NaoPago = Total_NaoPago + 1
            guia_resultados.cell(2, 11).value = ocorrências
            guia_resultados.cell(2, 11).alignment = Alignment(horizontal='center', vertical='center')

            Total_Vl_Nao_Pago += Total
            guia_resultados.cell(4, 11).value = Total_Vl_Nao_Pago

            Total_Juros_N += juros
            guia_resultados.cell(3, 11).value = Total_Juros_N
            guia_resultados.cell(3, 11).number_format = '#,##0.00'
            guia_resultados.cell(3, 11).alignment = Alignment(horizontal='center', vertical='center')
    elif situacao == "Confirmado":
        if juros == 0:
            mora = 0
            desc = 0.035
            Total = valor - (valor * desc)
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 13).value = desc
            guia_contas_receber.cell(linha, 17).value = Total

            ocorrências = Total_Confirmado = Total_Confirmado + 1
            guia_resultados.cell(2, 12).value = ocorrências
            guia_resultados.cell(2, 12).alignment = Alignment(horizontal='center', vertical='center')

            Total_Vl_Confirmado += Total
            guia_resultados.cell(4, 12).value = Total_Vl_Confirmado
            guia_resultados.cell(4, 12).number_format = '#,##0.00'
            guia_resultados.cell(4, 12).alignment = Alignment(horizontal='center', vertical='center')
        else:
            mora = 0
            desc = 0.035
            Total = valor + juros
            guia_contas_receber.cell(linha, 15).value = mora
            guia_contas_receber.cell(linha, 13).value = desc
            guia_contas_receber.cell(linha, 17).value = Total

            ocorrências = Total_Confirmado = Total_Confirmado + 1
            guia_resultados.cell(2, 12).value = ocorrências
            guia_resultados.cell(2, 12).alignment = Alignment(horizontal='center', vertical='center')

            Total_Vl_Confirmado += Total
            guia_resultados.cell(4, 12).value = Total_Vl_Confirmado
            guia_resultados.cell(4, 12).number_format = '#,##0.00'
            guia_resultados.cell(4, 12).alignment = Alignment(horizontal='center', vertical='center')

            Total_Juros_C += juros
            guia_resultados.cell(3, 12).value = Total_Juros_C
            guia_resultados.cell(3, 12).number_format = '#,##0.00'
            guia_resultados.cell(3, 12).alignment = Alignment(horizontal='center', vertical='center')

# O OBJETIVO É TOTALIZAR POR MÊS / DIA TODAS AS FATURAS VENCIDAS
# CUJO STATUS SEJA 'CONFIRMADO' E 'ATRASADO' NA TABELA CALENDÁRIO
# DA GUIA 'ANÁLISE' E, TOTALIZAR O VALOR DE DESCONTO DADO ÀS FATURAS
# COM STATUS 'CONFIRMADO'.


arquivo.save("CUSTO DA OPERACAO 3.xlsx")
