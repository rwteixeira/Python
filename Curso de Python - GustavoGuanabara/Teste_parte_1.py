# NESSA ETAPA, O SISTEMA IRÁ IDENTIFICAR OS MESES DO FATURAMENTO
# E OS COLOCARÁ EM ORDEM CRESCENTE NA GUIA 'ANALISE'.

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo = load_workbook("CUSTO DA OPERACAO 2.xlsx")
guia_contas_receber = arquivo["Python"]
guia_resultados = arquivo["Analise"]

meses_numeros = []
guia_resultados.cell(1, 1).value = "Meses:"
guia_resultados.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
guia_resultados.cell(1, 1).font = Font(name='Consolas', size=11, bold=True)

for cell in guia_contas_receber['I'][4:]:
    if cell.value is not None:
        meses_numeros.append(cell.value)

meses_numeros_unicos = sorted(set(meses_numeros))
Resultado = meses_numeros_unicos

if "Analise" in arquivo.sheetnames:
    novo_ws = arquivo["Analise"]
else:
    novo_ws = arquivo.create_sheet(title="Analise")

for idx, mes in enumerate(Resultado, start=2):
    guia_resultados.cell(idx, 1).alignment = Alignment(horizontal='center', vertical='center')
    novo_ws.cell(row=idx, column=1, value=mes)




arquivo.save("CUSTO DA OPERACAO 3.xlsx")
print("Resultado:", Resultado)
