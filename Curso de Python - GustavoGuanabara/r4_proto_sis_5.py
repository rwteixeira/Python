from openpyxl import load_workbook

# Carregar o arquivo
file_path = "MAPA_APP_dev.xlsx"  # Atualize com o caminho correto
workbook = load_workbook(file_path)
sheet = workbook["MOBILIZADO"]  # Seleciona a aba correta

# Obter os cabeçalhos da linha 3
header_row = next(sheet.iter_rows(min_row=3, max_row=3, values_only=True))

# Identificar as colunas "PREFIXO" e "ORDEM"
try:
    col_prefixo = header_row.index("PREFIXO")  # Índice da coluna PREFIXO
    col_ordem = header_row.index("ORDEM")  # Índice da coluna ORDEM

    # Criar os códigos concatenando PREFIXO + ORDEM a partir da linha 4
    codigos = [
        f"{row[col_prefixo]}-{row[col_ordem]}"
        for row in sheet.iter_rows(min_row=4, values_only=True)
        if row[col_prefixo] and row[col_ordem]  # Ignora células vazias
    ]

    # print(codigos[:])  # Exibir os primeiros 10 códigos
    # print(len(codigos))

    for i, c in enumerate(codigos):
        print(f'{codigos[i]}')


except ValueError as e:
    print(f"Erro: {e}")
