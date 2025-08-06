import logging
from openpyxl import load_workbook
from openpyxl import Workbook
import pyautogui


def pasta_de_trabalho():
    arquivo = pyautogui.confirm(text='Digite o nome do arquivo:', )
    nome_arquivo = str(input('Defina o nome do arquivo: '))
    wb = Workbook()
    ws = wb.active
    ws1 = wb.create_sheet(title="Planilha1")
    ws2 = wb.create_sheet(title="Planilha2")
    wb.save(nome_arquivo + '.xlsx')

# Definir o nome da Pasta de trabalho do Excel
nome_do_arquivo = str(input('Digite o nome da pasta de trabalho: '))

# Configuração do logger
logging.basicConfig(
    filename='Excel Python 1.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
