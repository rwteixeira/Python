from openpyxl import load_workbook, Workbook

wb: Workbook = load_workbook("CUSTO DA OPERACAO 2.xlsx")
ws = wb["Python"]
meses_numeros = []

for cell in ws['I'][4:]:
    if cell.value is not None:
        meses_numeros.append(cell.value)

meses_numeros_unicos = sorted(set(meses_numeros))
Resultado = meses_numeros_unicos

if "Analise" in wb.sheetnames:
    novo_ws = wb["Analise"]
else:
    novo_ws = wb.create_sheet(title="Analise")

for idx, mes in enumerate(Resultado, start=2):
    novo_ws.cell(row=idx, column=1, value=mes)

wb.save("CUSTO DA OPERACAO 3.xlsx")
print("Resultado:", Resultado)
